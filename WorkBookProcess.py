#工作簿的各种操作
#
import openpyxl

wb=openpyxl.load_workbook("top250movie.xlsx")#打开指定路径的工作簿
ws=wb["Sheet"]#获得指定名称的工作表
#==================
#获取A2单元格
#c=ws['A2']
#=========================
#创建工作表
#wb.create_sheet(index=0,title="wonkmy demo")
#============================
#移除工作表
#wb.remove(wb["wonkmy demo"])
#===================
#获取c单元格的内容
#c.value 
#===================
#获取c单元格向下偏移2个单位、第0列的内容
#c.offset(2,0)
#===================
 #获取序号第496个单元格的坐标
#openpyxl.cell.cell.get_column_letter(496)
#===================
# 获取坐标“JB”单元格的序号
#openpyxl.cell.cell.column_index_from_string("JB") 
#===================
# 访问A2到B10之间的所有单元格
#for each_movie in ws['A2':'B10']:
#    for each_cell in each_movie:
#        print(each_cell.value,end=' ')
#    print('\n')
#===============
#拷贝指定工作表
#new=wb.copy_worksheet(ws)
#wb.save("top250movie1.xlsx")
#================
#将指定的单元格内容修改成自己指定的内容
#updateData= {"阿甘正传": "9.8","这个杀手不太冷": "9.6","肖申克的救赎": "9.7"}
#for each_row in ws.rows:
#    if each_row[0].value in updateData:
#        each_row[1].value=updateData[each_row[0].value]
#wb.save("top250movie.xlsx")
#====================
#print(wb.sheetnames)