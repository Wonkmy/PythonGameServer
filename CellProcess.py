from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import GradientFill
from openpyxl.styles import Border,Side
from openpyxl.styles import Alignment
from openpyxl.styles import NamedStyle

#创建工作簿、工作表和单元格
wb=Workbook()
ws=wb.active

b2=ws['B2']
b2.value="wonkmy"
b3=ws['B3']
b3.value="wonkmy"
#=========================================================================================
#设置b2单元格字体
#bold_red_font = Font(bold=True,color="FF0000")
#b2.font=bold_red_font

#设置b3单元格字体
#bold_blue_font = Font(size=16,italic=True,strike=True,color="0000FF")
#b3.font=bold_blue_font
#=========================================================================================
#填充单元格
#yellowfill=PatternFill(fill_type="solid",fgColor="FFFF00")#纯色填充
#b2.fill=yellowfill

#red2grenn_fill=GradientFill(type="linear",stop=("FF0000","00FF00"))#线性填充
#b3.fill=red2grenn_fill
#=========================================================================================
#设置边框
#thin_side=Side(border_style="thin",color="000000")
#double_side=Side(border_style="double",color="FF0000")

#b2.border=Border(diagonal=thin_side,diagonalUp=True,diagonalDown=True)
#b3.border=Border(left=double_side,top=double_side,right=double_side,bottom=double_side)
#=========================================================================================
#文本对齐
#ws.merge_cells('A1:C2')
#ws['A1'].value="wonkmygame.com"

#centeraligment=Alignment(horizontal="center",vertical="center")
#ws['A1'].alignment=centeraligment
#=========================================================================================
#命名样式  即:将前面一系列的样式设置成一个整体模板,然后应用到某个单元格里
#hightlight=NamedStyle(name="hightlight")
#hightlight.font=Font(bold=True,size=20)
#hightlight.alignment=Alignment(horizontal="center",vertical="center")
#wb.add_named_style(hightlight)

#ws['A1'].style=hightlight
#ws['B5'].value="love"
#ws['B5'].style=hightlight
#=========================================================================================
wb.save("1.xlsx")