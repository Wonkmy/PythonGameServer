import openpyxl
wb=openpyxl.load_workbook("WorkTest.xlsx")
ws=wb.active
serch_content="待开发"

for each_L in ws['L']:
    each_L.value ="已完成"
#wb=openpyxl.Workbook()
#===============================
#修改指定工作表的标签颜色
#ws1=wb.create_sheet(title="wonkmy")
#ws2=wb.create_sheet(title="game")
#ws3=wb.create_sheet(title="wonkmygame")

#ws1.sheet_properties.tabColor="FF0000"
#ws2.sheet_properties.tabColor="00FF00"
#ws3.sheet_properties.tabColor="0000FF"

#
#================================
#更改指定行(第2行)的行高和指定列(第C列)的列宽
#ws1.row_dimensions[2].height=100
#ws1.column_dimensions['C'].width=100
#=============================
#合并和拆分单元格
#ws1.merge_cells("A1:C3")
#ws1.unmerge_cells("A1:C3")
#==============================
#冻结和解冻单元格
#ws1.freeze_panes='B3'#B3的意思是，以B3单元格为基准，左边和上边的单元格冻结，自身不冻结
#ws1.freeze_panes='A1'#B3的意思是，以A1单元格为基准，左边和上边的单元格冻结，自身不冻结,即可实现解冻
wb.save("WorkTest.xlsx")
