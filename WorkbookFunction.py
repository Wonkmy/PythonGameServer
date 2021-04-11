from openpyxl.utils import FORMULAE
from openpyxl import load_workbook
wb=load_workbook("test.xlsx")

ws=wb['Sheet']
# for row in ws.iter_rows(min_col=2,min_row=2,max_col=5,max_row=5):
#     ws[row[3].coordinate]="=SUM(%s:%s)" %(row[0].coordinate,row[2].coordinate)

# for row in ws.iter_rows(min_col=2,min_row=2,max_col=6,max_row=5):
#     ws[row[4].coordinate]='=IF(%s>250,"A","B")' % (row[3].coordinate)


wb.save("test.xlsx")